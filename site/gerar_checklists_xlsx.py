# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl"]
# ///
"""
Gera os checklists em planilha (.xlsx) editável para download, a partir do
conteúdo Markdown dos checklists do site (fonte única).

Uso:
    python site/gerar_checklists_xlsx.py

Saída: site/assets/downloads/*.xlsx  → publicados no build (copytree de assets/).
"""

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent.parent
CONTEUDO = BASE_DIR / "conteudo"
OUT_DIR = Path(__file__).resolve().parent / "assets" / "downloads"
OUT_DIR.mkdir(parents=True, exist_ok=True)

ROXO = "7A10CC"
ROXO_CLARO = "EFE2FB"
CINZA_CLARO = "F2F2F2"
BRANCO = "FFFFFF"

PRIORIDADES = ("Essencial", "Crítico", "Importante", "Diferencial")

ITEM_RE = re.compile(r"^(?P<indent>\s*)- \[[ xX]?\]\s*(?P<texto>.*)$")
H2_RE = re.compile(r"^##\s+(?P<t>.+)$")
H3_RE = re.compile(r"^###\s+(?P<t>.+)$")
BOLD_LEADIN_RE = re.compile(r"^\*\*(?P<t>.+?):\*\*\s*$")


def limpa_md(texto: str) -> tuple[str, str]:
    """Remove marcação Markdown e extrai a prioridade textual, se houver."""
    prioridade = ""
    m = re.match(r"^\((?P<p>[^)]+)\)\s*(?P<resto>.*)$", texto)
    if m and m.group("p") in PRIORIDADES:
        prioridade = m.group("p")
        texto = m.group("resto")
    texto = texto.replace("**", "").replace("*", "")
    texto = re.sub(r"`([^`]*)`", r"\1", texto)
    return texto.strip(), prioridade


def parse_checklist(md_path: Path):
    """Retorna lista de seções: [(nome_secao, [(categoria, item, prioridade, nivel), ...]), ...]."""
    secoes = []
    secao_atual = None
    categoria = ""
    grupo = ""

    for raw in md_path.read_text(encoding="utf-8").splitlines():
        line = raw.rstrip()
        if not line:
            continue

        h2 = H2_RE.match(line)
        if h2:
            titulo = h2.group("t").strip()
            # Ignora seções que não fazem parte do checklist propriamente dito
            secao_atual = (titulo, [])
            secoes.append(secao_atual)
            categoria = ""
            grupo = ""
            continue

        h3 = H3_RE.match(line)
        if h3:
            categoria = h3.group("t").strip()
            grupo = ""
            if secao_atual is None:
                secao_atual = ("Checklist", [])
                secoes.append(secao_atual)
            continue

        bold = BOLD_LEADIN_RE.match(line)
        if bold:
            grupo = bold.group("t").strip()
            continue

        item = ITEM_RE.match(line)
        if item:
            if secao_atual is None:
                secao_atual = ("Checklist", [])
                secoes.append(secao_atual)
            nivel = len(item.group("indent")) // 2
            texto, prioridade = limpa_md(item.group("texto"))
            if not texto:
                continue
            cat = " — ".join([p for p in (categoria, grupo) if p])
            secao_atual[1].append((cat, texto, prioridade, nivel))

    # Mantém apenas seções que têm itens de checklist
    return [(nome, itens) for nome, itens in secoes if itens]


def _nome_aba(titulo: str) -> str:
    t = re.sub(r"^FASE\s*\d+:\s*", "", titulo, flags=re.IGNORECASE)
    t = re.sub(r"[\[\]\:\*\?/\\]", "", t).strip()
    return (t[:31] or "Checklist")


def escreve_planilha(secoes, titulo_doc: str, out_path: Path):
    wb = Workbook()
    wb.remove(wb.active)

    thin = Side(style="thin", color="DDDDDD")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor=ROXO)
    header_font = Font(color=BRANCO, bold=True, size=11)
    cat_font = Font(bold=True, color=ROXO)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center")

    headers = ["OK", "Categoria", "Item", "Prioridade", "Observações"]
    larguras = [6, 32, 70, 14, 30]

    for nome_secao, itens in secoes:
        ws = wb.create_sheet(_nome_aba(nome_secao))

        # Título da aba
        ws.merge_cells("A1:E1")
        c = ws["A1"]
        c.value = nome_secao
        c.font = Font(bold=True, size=13, color=ROXO)
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[1].height = 22

        # Cabeçalho da tabela
        for col, (h, w) in enumerate(zip(headers, larguras), start=1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = borda
            ws.column_dimensions[get_column_letter(col)].width = w

        r = 3
        for cat, item, prioridade, nivel in itens:
            prefixo = "    ↳ " if nivel > 0 else ""
            ws.cell(row=r, column=1, value="").border = borda
            cc = ws.cell(row=r, column=2, value=cat)
            cc.alignment = wrap
            cc.font = cat_font
            cc.border = borda
            ic = ws.cell(row=r, column=3, value=prefixo + item)
            ic.alignment = wrap
            ic.border = borda
            pc = ws.cell(row=r, column=4, value=prioridade)
            pc.alignment = center
            pc.border = borda
            if prioridade in ("Essencial", "Crítico"):
                pc.font = Font(bold=True, color="C0392B")
            ws.cell(row=r, column=5, value="").border = borda
            if r % 2 == 1:
                for col in range(1, 6):
                    if not ws.cell(row=r, column=col).fill.fgColor.rgb or \
                       ws.cell(row=r, column=col).fill.patternType is None:
                        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=CINZA_CLARO)
            r += 1

        ws.freeze_panes = "A3"

    # Aba de instruções como primeira
    info = wb.create_sheet("Como usar", 0)
    info["A1"] = titulo_doc
    info["A1"].font = Font(bold=True, size=14, color=ROXO)
    instrucoes = [
        "",
        "Como usar este checklist:",
        "• Marque a coluna OK (com um X) à medida que cada item for cumprido.",
        "• Use a coluna Observações para registrar evidências, responsáveis e prazos.",
        "• Prioridade \"Essencial\"/\"Crítico\": o mínimo para uma contratação responsável.",
        "• Cada aba corresponde a uma fase ou bloco do checklist.",
        "",
        "Material da Aliança de IA para a Educação, em parceria com o Instituto Jataí.",
        "Derivado da publicação \"Contratação Pública de Soluções de IA na Educação\".",
    ]
    for i, txt in enumerate(instrucoes, start=2):
        info.cell(row=i, column=1, value=txt)
    info.column_dimensions["A"].width = 100

    wb.save(out_path)
    total = sum(len(itens) for _, itens in secoes)
    print(f"  ✓ {out_path.name} ({len(secoes)} abas, {total} itens)")


def main():
    print("Gerando checklists (.xlsx)…")
    # Jornada completa dos gestores (3 fases)
    secoes = parse_checklist(CONTEUDO / "recursos" / "checklist-gestores.md")
    escreve_planilha(
        secoes,
        "Checklist da Jornada de Contratação de IA — Gestores",
        OUT_DIR / "checklist-jornada-gestores.xlsx",
    )
    # Checklist das EdTechs
    secoes = parse_checklist(CONTEUDO / "edtechs" / "checklist-edtechs.md")
    escreve_planilha(
        secoes,
        "Checklist de Prontidão para o Setor Público — EdTechs",
        OUT_DIR / "checklist-edtechs.xlsx",
    )
    print("Concluído.")


if __name__ == "__main__":
    main()
